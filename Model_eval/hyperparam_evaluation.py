import numpy as np
from sklearn.metrics import mean_squared_error, mean_absolute_error
import pandas as pd
import matplotlib.pyplot as plt
import time

def smape(y_true, y_pred):
    y_true = np.asarray(y_true, dtype = float)
    y_pred = np.asarray(y_pred, dtype = float)

    denominator = np.abs(y_true) + np.abs(y_pred)
    mask = denominator != 0

    values = np.zeros_like(y_true, dtype = float)
    values[mask] = 2 * np.abs(y_pred[mask] - y_true[mask]) / denominator[mask]

    return 100 * np.mean(values)

def cal_errors(original_df):
    list_actuals = original_df["DKPrice"].astype(float).values
    list_predictions = original_df["Prediction"].astype(float).values
    list_errors = np.abs(list_actuals - list_predictions)

    return list_errors

def error_dateframe(original_df, list_errors):
    new_df = original_df[["Time"]]

    new_df["Errors"] = list_errors

    return new_df

def generate_summer_df(error_df):
    error_df["Time"] = pd.to_datetime(error_df["Time"], format = '%Y-%m-%d %H:%M:%S')

    summer_df = error_df.loc[(error_df["Time"] >= "2024-04-01 00:00:00")
                             & (error_df["Time"] <= "2024-09-30 23:00:00")]
    
    return summer_df

def generate_winter_df(error_df):
    error_df["Time"] = pd.to_datetime(error_df["Time"], format = '%Y-%m-%d %H:%M:%S')

    winter_df = error_df.loc[(error_df["Time"] < "2024-04-01 00:00:00")
                             | (error_df["Time"] > "2024-09-30 23:00:00")]
    
    return winter_df

def calc_metrics(summer_df, winter_df, original_df):
    list_actuals = original_df["DKPrice"].astype(float).values
    list_predictions = original_df["Prediction"].astype(float).values
    list_errors = np.abs(list_actuals - list_predictions)
    summer_list = summer_df["Errors"].astype(float).values
    winter_list = winter_df["Errors"].astype(float).values

    overall_smape = smape(list_actuals, list_predictions)
    overall_rmse = np.sqrt(mean_squared_error(list_actuals, list_predictions))
    overall_mae = mean_absolute_error(list_actuals, list_predictions)
    max_error = max(list_errors)
    min_error = min(list_errors)
    q1_metric = np.percentile(list_errors, 25)
    q2_metric = np.percentile(list_errors, 50)
    q3_metric = np.percentile(list_errors, 75)  
    mean_error = sum(list_errors) / len(list_errors)
    var_error = sum((i - mean_error) ** 2 for i in list_errors) / len(list_errors)
    mean_summer_error = sum(summer_list) / len(summer_list)
    mean_winter_error = sum(winter_list) / len(winter_list)

    return (overall_smape, overall_rmse, overall_mae, max_error, min_error,
            q1_metric, q2_metric, q3_metric, mean_error, var_error,
            mean_summer_error, mean_winter_error)


METRIC_COLS = [
    "smape", "rmse", "mae", "max_error", "min_error",
    "q1", "q2_median", "q3", "mean_error", "var_error",
    "mean_summer_error", "mean_winter_error",
]


def load_prediction_files(folder_path):
    """
    Scan folder_path for CSV files containing 'prediction' in the filename.
    Returns two dicts: dk1_files and dk2_files, each mapping filename -> DataFrame.
    Files must contain 'DK1' or 'DK2' in the name to be included.
    """
    import os

    dk1_files, dk2_files = {}, {}

    for fname in os.listdir(folder_path):
        if "prediction" not in fname.lower():
            continue
        if not fname.endswith(".csv"):
            continue

        fpath = os.path.join(folder_path, fname)
        df = pd.read_csv(fpath, decimal= ",")

        if "DK1" in fname:
            dk1_files[fname] = df
        elif "DK2" in fname:
            dk2_files[fname] = df

    return dk1_files, dk2_files


def _evaluate_single_file(fname, df):
    """Run the full metric pipeline for one prediction file. Returns a dict."""
    list_errors = cal_errors(df)
    error_df = error_dateframe(df, list_errors)
    summer_df = generate_summer_df(error_df.copy())
    winter_df = generate_winter_df(error_df.copy())

    (overall_smape, overall_rmse, overall_mae, max_error, min_error,
     q1_metric, q2_metric, q3_metric, mean_error, var_error,
     mean_summer_error, mean_winter_error) = calc_metrics(summer_df, winter_df, df)

    return {
        "filename": fname,
        "smape": overall_smape,
        "rmse": overall_rmse,
        "mae": overall_mae,
        "max_error": max_error,
        "min_error": min_error,
        "q1": q1_metric,
        "q2_median": q2_metric,
        "q3": q3_metric,
        "mean_error": mean_error,
        "var_error": var_error,
        "mean_summer_error": mean_summer_error,
        "mean_winter_error": mean_winter_error,
    }


def evaluate_all(folder_path):
    """
    Load all prediction files, compute metrics for each, and return two DataFrames:
    one for DK1 and one for DK2, each with one row per hyperparameter combination.
    """
    dk1_files, dk2_files = load_prediction_files(folder_path)

    dk1_rows = [_evaluate_single_file(fname, df) for fname, df in dk1_files.items()]
    dk2_rows = [_evaluate_single_file(fname, df) for fname, df in dk2_files.items()]

    dk1_df = pd.DataFrame(dk1_rows)
    dk2_df = pd.DataFrame(dk2_rows)

    return dk1_df, dk2_df


def compute_scores(results_df):
    """
    Add a composite score column to a results DataFrame using min-max normalisation
    across all 12 metrics (lower raw value = better for every metric).
    Score of 0.0 = best possible, 1.0 = worst possible.
    Returns a copy of the DataFrame sorted by score ascending (best first).
    """
    df = results_df.copy()

    normalised = pd.DataFrame(index=df.index)
    for col in METRIC_COLS:
        col_min = df[col].min()
        col_max = df[col].max()
        denom = col_max - col_min
        if denom == 0:
            # All combos identical on this metric — contributes nothing to ranking
            normalised[col] = 0.0
        else:
            normalised[col] = (df[col] - col_min) / denom

    df["score"] = normalised[METRIC_COLS].mean(axis=1)
    df["rank"] = df["score"].rank(method="min").astype(int)
    df = df.sort_values("score").reset_index(drop=True)

    return df


def save_results(dk1_df, dk2_df, output_path):
    """
    Score and rank both zone DataFrames, then save them as two sheets
    (DK1, DK2) in an Excel file at output_path.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    dk1_scored = compute_scores(dk1_df)
    dk2_scored = compute_scores(dk2_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dk1_scored.to_excel(writer, sheet_name="DK1", index=False)
        dk2_scored.to_excel(writer, sheet_name="DK2", index=False)

        for zone, df in [("DK1", dk1_scored), ("DK2", dk2_scored)]:
            ws = writer.sheets[zone]

            # Header formatting
            header_fill = PatternFill("solid", start_color="2F4F8F", end_color="2F4F8F")
            header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Score and rank column highlighting (last two columns)
            score_fill = PatternFill("solid", start_color="E8F4E8", end_color="E8F4E8")
            score_col_indices = [
                df.columns.get_loc("score") + 1,
                df.columns.get_loc("rank") + 1,
            ]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in score_col_indices:
                    row[col_idx - 1].fill = score_fill

            # Top-3 rows highlighted in gold
            gold_fill = PatternFill("solid", start_color="FFD700", end_color="FFD700")
            rank_col_idx = df.columns.get_loc("rank") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                rank_val = row[rank_col_idx - 1].value
                if rank_val is not None and rank_val <= 3:
                    for cell in row:
                        cell.fill = gold_fill

            # Auto-size columns
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 12)

            # Freeze header row
            ws.freeze_panes = "A2"

    print(f"Results saved to {output_path}")